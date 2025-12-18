#!/usr/bin/env python3
"""
Creates a sample groups.xlsx file for the drchrono application.

Format:
- A1: Trial: <trial_name>
- A2: Groups
- A3 onwards: Group names (one per row)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Groups"

# Add trial name in A1
ws['A1'] = "Trial: DR EXTREMO 2025"
ws['A1'].font = Font(bold=True, size=14)

# Add header in A2
ws['A2'] = "Groups"
ws['A2'].font = Font(bold=True, size=12)
ws['A2'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# Add sample groups
groups = ["Pro", "Sport", "Tourism", "Elite", "Amateur"]
for idx, group in enumerate(groups, start=3):
    ws[f'A{idx}'] = group

# Adjust column width
ws.column_dimensions['A'].width = 20

# Save file
output_path = "/home/runner/work/drchrono/drchrono/groups.xlsx"
wb.save(output_path)
print(f"Sample XLSX file created: {output_path}")
print(f"Trial: DR EXTREMO 2025")
print(f"Groups: {', '.join(groups)}")
